"""数据层：拉取 AkShare 三大表 + 计算固定指标/杜邦/同比，生成对比结构，导出 Excel。
财报接口：stock_*_sheet_by_report_em（行=报告期，列=英文科目代码，含 *_YOY 同比列）。
行业/同行接口：离线缓存 peers_cache.json（申万二级行业，由 build_peers_cache.py 一次性生成），
运行时零网络依赖，可在被拦截 Eastmoney 行情接口的网络下正常使用。
"""
import akshare as ak
import pandas as pd
import warnings
import time
import math
import os
import json
import re
import sqlite3
from datetime import datetime

warnings.filterwarnings('ignore')

METADATA_COLS = {
    'SECUCODE', 'SECURITY_CODE', 'SECURITY_NAME_ABBR', 'ORG_CODE', 'ORG_TYPE',
    'REPORT_DATE', 'REPORT_TYPE', 'REPORT_DATE_NAME', 'SECURITY_TYPE_CODE',
    'NOTICE_DATE', 'UPDATE_DATE', 'CURRENCY', 'OPINION_TYPE', 'OSOPINION_TYPE', 'LISTING_STATE',
}

# 英文科目代码 -> 中文名（覆盖主要科目；未列出的保留英文代码）
NAME_MAP = {
    # 资产负债
    'MONETARYFUNDS': '货币资金', 'TRADABLE_FINASSET': '交易性金融资产', 'NOTE_RECE': '应收票据',
    'ACCOUNTS_RECE': '应收账款', 'FINANCE_RECE': '应收利息', 'DIVIDEND_RECE': '应收股利',
    'PREPAYMENT': '预付款项', 'OTHER_RECE': '其他应收款', 'INVENTORY': '存货',
    'CONTRACT_ASSET': '合同资产', 'OTHER_CURRENT_ASSET': '其他流动资产', 'TOTAL_CURRENT_ASSETS': '流动资产合计',
    'CIP': '在建工程', 'FIXED_ASSET': '固定资产', 'CONSUMPTIVE_BIOLOGICAL_ASSET': '消耗性生物资产',
    'PRODUCTIVE_BIOLOGY_ASSET': '生产性生物资产', 'INVEST_REALESTATE': '投资性房地产',
    'INTANGIBLE_ASSET': '无形资产', 'DEVELOP_EXPENSE': '开发支出', 'GOODWILL': '商誉',
    'LONG_EQUITY_INVEST': '长期股权投资', 'OTHER_EQUITY_INVEST': '其他权益工具投资',
    'HOLD_MATURITY_INVEST': '持有至到期投资', 'LONG_RECE': '长期应收款', 'DEFER_TAX_ASSET': '递延所得税资产',
    'OTHER_NONCURRENT_ASSET': '其他非流动资产', 'NONCURRENT_ASSET_1YEAR': '一年内到期的非流动资产',
    'TOTAL_NONCURRENT_ASSETS': '非流动资产合计', 'TOTAL_ASSETS': '资产总计',
    'SHORT_LOAN': '短期借款', 'NOTE_PAYABLE': '应付票据', 'ACCOUNTS_PAYABLE': '应付账款',
    'ADVANCE_RECEIVABLES': '预收款项', 'CONTRACT_LIAB': '合同负债', 'STAFF_SALARY_PAYABLE': '应付职工薪酬',
    'TAX_PAYABLE': '应交税费', 'INTEREST_PAYABLE': '应付利息', 'DIVIDEND_PAYABLE': '应付股利',
    'OTHER_PAYABLE': '其他应付款', 'DEFER_INCOME': '递延收益', 'SHORT_FIN_PAYABLE': '应付短期融资款',
    'SHORT_BOND_PAYABLE': '应付短期债券', 'CURRENT_LIAB_OTHER': '其他流动负债', 'TOTAL_CURRENT_LIAB': '流动负债合计',
    'LONG_LOAN': '长期借款', 'BOND_PAYABLE': '应付债券', 'LONG_PAYABLE': '长期应付款',
    'SPECIAL_PAYABLE': '专项应付款', 'DEFER_TAX_LIAB': '递延所得税负债', 'OTHER_NONCURRENT_LIAB': '其他非流动负债',
    'NONCURRENT_LIAB_1YEAR': '一年内到期的非流动负债', 'TOTAL_NONCURRENT_LIAB': '非流动负债合计',
    'TOTAL_LIABILITIES': '负债合计', 'SHARE_CAPITAL': '股本', 'CAPITAL_RESERVE': '资本公积',
    'SURPLUS_RESERVE': '盈余公积', 'UNASSIGN_RPOFIT': '未分配利润', 'OTHER_COMPRE_INCOME': '其他综合收益',
    'SPECIAL_RESERVE': '专项储备', 'TOTAL_PARENT_EQUITY': '归属于母公司股东权益合计',
    'MINORITY_EQUITY': '少数股东权益', 'TOTAL_EQUITY': '所有者权益合计', 'TOTAL_LIAB_EQUITY': '负债和所有者权益总计',
    # 利润
    'TOTAL_OPERATE_INCOME': '营业总收入', 'OPERATE_INCOME': '营业收入', 'TOTAL_OPERATE_COST': '营业总成本',
    'OPERATE_COST': '营业成本', 'OPERATE_TAX_ADD': '税金及附加', 'SALE_EXPENSE': '销售费用',
    'MANAGE_EXPENSE': '管理费用', 'RESEARCH_EXPENSE': '研发费用', 'FINANCE_EXPENSE': '财务费用',
    'FE_INTEREST_EXPENSE': '利息支出', 'FE_INTEREST_INCOME': '利息收入', 'FAIRVALUE_CHANGE_INCOME': '公允价值变动收益',
    'INVEST_INCOME': '投资收益', 'ASSET_IMPAIRMENT_LOSS': '资产减值损失', 'CREDIT_IMPAIRMENT_LOSS': '信用减值损失',
    'OTHER_INCOME': '其他收益', 'OPERATE_PROFIT': '营业利润', 'NONBUSINESS_INCOME': '营业外收入',
    'NONBUSINESS_EXPENSE': '营业外支出', 'TOTAL_PROFIT': '利润总额', 'INCOME_TAX': '所得税费用',
    'NETPROFIT': '净利润', 'PARENT_NETPROFIT': '归母净利润', 'DEDUCT_PARENT_NETPROFIT': '扣非归母净利润',
    'MINORITY_INTEREST': '少数股东损益', 'BASIC_EPS': '基本每股收益', 'DILUTED_EPS': '稀释每股收益',
    'TOTAL_COMPRE_INCOME': '综合收益总额',
    # 现金流
    'SALES_SERVICES': '销售商品提供劳务收到的现金', 'RECEIVE_TAX_REFUND': '收到的税费返还',
    'RECEIVE_OTHER_OPERATE': '收到其他与经营活动有关的现金', 'TOTAL_OPERATE_INFLOW': '经营活动现金流入小计',
    'BUY_SERVICES': '购买商品接受劳务支付的现金', 'PAY_STAFF_CASH': '支付给职工以及为职工支付的现金',
    'PAY_ALL_TAX': '支付的各项税费', 'PAY_OTHER_OPERATE': '支付其他与经营活动有关的现金',
    'TOTAL_OPERATE_OUTFLOW': '经营活动现金流出小计', 'NETCASH_OPERATE': '经营活动现金流量净额',
    'WITHDRAW_INVEST': '收回投资收到的现金', 'RECEIVE_INVEST_INCOME': '取得投资收益收到的现金',
    'DISPOSAL_LONG_ASSET': '处置固定资产等收回的现金净额', 'TOTAL_INVEST_INFLOW': '投资活动现金流入小计',
    'CONSTRUCT_LONG_ASSET': '购建固定资产等支付的现金', 'INVEST_PAY_CASH': '投资支付的现金',
    'OBTAIN_SUBSIDIARY_OTHER': '取得子公司支付的现金净额', 'TOTAL_INVEST_OUTFLOW': '投资活动现金流出小计',
    'NETCASH_INVEST': '投资活动现金流量净额', 'RECEIVE_LOAN_CASH': '取得借款收到的现金',
    'ISSUE_BOND': '发行债券收到的现金', 'RECEIVE_OTHER_FINANCE': '收到其他与筹资活动有关的现金',
    'TOTAL_FINANCE_INFLOW': '筹资活动现金流入小计', 'PAY_DEBT_CASH': '偿还债务支付的现金',
    'ASSIGN_DIVIDEND_PORFIT': '分配股利利润或偿付利息支付的现金', 'TOTAL_FINANCE_OUTFLOW': '筹资活动现金流出小计',
    'NETCASH_FINANCE': '筹资活动现金流量净额', 'CCE_ADD': '现金及现金等价物净增加额',
    'END_CCE': '期末现金及现金等价物余额', 'BEGIN_CCE': '期初现金及现金等价物余额',
    # 补充：财报标准科目（确保不出现英文）
    'TRADABLE_FINASSET': '交易性金融资产', 'TRADE_FINASSET': '交易性金融资产',
    'DERIVE_FINASSET': '衍生金融资产', 'NOTE_ACCOUNTS_RECE': '应收款项融资',
    'INTEREST_RECE': '应收利息', 'HOLDSALE_ASSET': '持有待售资产',
    'LOAN_ADVANCE': '发放贷款及垫款', 'AMORTIZE_COST_FINASSET': '债权投资',
    'FVTOCI_FINASSET': '其他债权投资', 'FVTPL_FINASSET': '其他非流动金融资产',
    'OIL_GAS_ASSET': '油气资产', 'USERIGHT_ASSET': '使用权资产',
    'LONG_PREPAID_EXPENSE': '长期待摊费用', 'LOAN_PBC': '向中央银行借款',
    'BORROW_FUND': '拆入资金', 'ACCEPT_DEPOSIT_INTERBANK': '吸收存款及同业存放',
    'TRADE_FINLIAB': '交易性金融负债', 'DERIVE_FINLIAB': '衍生金融负债',
    'HOLDSALE_LIAB': '持有待售负债', 'LEASE_LIAB': '租赁负债',
    'LONG_STAFFSALARY_PAYABLE': '长期应付职工薪酬', 'PREDICT_LIAB': '预计负债',
    'OTHER_EQUITY_TOOL': '其他权益工具', 'TREASURY_SHARES': '库存股',
    'GENERAL_RISK_RESERVE': '一般风险准备',
    'INTEREST_INCOME': '利息收入', 'FEE_COMMISSION_INCOME': '手续费及佣金收入',
    'OTHER_BUSINESS_INCOME': '其他业务收入', 'INTEREST_EXPENSE': '利息支出',
    'FEE_COMMISSION_EXPENSE': '手续费及佣金支出',
    'RECEIVE_OTHER_INVEST': '收到其他与投资活动有关的现金',
    'PAY_OTHER_INVEST': '支付其他与投资活动有关的现金',
    'PAY_OTHER_FINANCE': '支付其他与筹资活动有关的现金',
}

# 港股财报（港交所口径，数值已为港币）：标准中文科目名 -> A 股同义科目代码
# 目的：复用现有 CANON_ORDER 行序 + _indicator_defs 全部固定指标，无需改动下游。
HK_ALIAS = {
    '总资产': 'TOTAL_ASSETS', '资产总计': 'TOTAL_ASSETS', '总负债': 'TOTAL_LIABILITIES',
    '负债总计': 'TOTAL_LIABILITIES', '流动资产合计': 'TOTAL_CURRENT_ASSETS',
    '非流动资产合计': 'TOTAL_NONCURRENT_ASSETS', '流动负债合计': 'TOTAL_CURRENT_LIAB',
    '非流动负债合计': 'TOTAL_NONCURRENT_LIAB', '总权益': 'TOTAL_EQUITY',
    '权益总额': 'TOTAL_EQUITY', '股东权益': 'TOTAL_EQUITY', '股东权益合计': 'TOTAL_EQUITY',
    '归属于母公司股东权益': 'TOTAL_PARENT_EQUITY', '归属于母公司权益': 'TOTAL_PARENT_EQUITY',
    '归属于公司权益持有人之权益': 'TOTAL_PARENT_EQUITY',
    '股本': 'SHARE_CAPITAL', '存货': 'INVENTORY', '应收帐款': 'ACCOUNTS_RECE',
    '应收账款': 'ACCOUNTS_RECE', '应收贸易账款': 'ACCOUNTS_RECE', '应付帐款': 'ACCOUNTS_PAYABLE',
    '应付贸易账款': 'ACCOUNTS_PAYABLE', '应付票据': 'NOTE_PAYABLE', '预付款项': 'PREPAYMENT',
    '固定资产': 'FIXED_ASSET', '在建工程': 'CIP', '无形资产': 'INTANGIBLE_ASSET',
    '投资物业': 'INVEST_REALESTATE', '投资性房地产': 'INVEST_REALESTATE',
    '现金及现金等价物': 'MONETARYFUNDS', '现金及银行结存': 'MONETARYFUNDS',
    '银行结余及现金': 'MONETARYFUNDS', '商誉': 'GOODWILL', '递延税项资产': 'DEFER_TAX_ASSET',
    '递延所得税资产': 'DEFER_TAX_ASSET', '递延税项负债': 'DEFER_TAX_LIAB',
    '递延所得税负债': 'DEFER_TAX_LIAB', '短期贷款': 'SHORT_LOAN', '短期借款': 'SHORT_LOAN',
    '长期贷款': 'LONG_LOAN', '长期借款': 'LONG_LOAN', '应付税项': 'TAX_PAYABLE',
    '应交税费': 'TAX_PAYABLE', '库存股': 'TREASURY_SHARES', '少数股东权益': 'MINORITY_EQUITY',
    '长期应付款': 'LONG_PAYABLE', '租赁负债': 'LEASE_LIAB', '预计负债': 'PREDICT_LIAB',
    '持有至到期投资': 'HOLD_MATURITY_INVEST', '联营公司权益': 'LONG_EQUITY_INVEST',
    '长期股权投资': 'LONG_EQUITY_INVEST', '于联营公司之投资': 'LONG_EQUITY_INVEST',
    '交易性金融资产': 'TRADABLE_FINASSET', '交易性金融资产(流动)': 'TRADABLE_FINASSET',
    '营业额': 'OPERATE_INCOME', '营业收入': 'OPERATE_INCOME', '营运支出': 'OPERATE_COST',
    '销售成本': 'OPERATE_COST', '经营支出总额': 'OPERATE_COST', '营业成本': 'OPERATE_COST',
    '经营溢利': 'OPERATE_PROFIT', '营业利润': 'OPERATE_PROFIT', '除税前溢利': 'TOTAL_PROFIT',
    '利润总额': 'TOTAL_PROFIT', '税项': 'INCOME_TAX', '所得税费用': 'INCOME_TAX',
    '股东应占溢利': 'PARENT_NETPROFIT', '本公司拥有人应占溢利': 'PARENT_NETPROFIT',
    '归属于母公司股东净利润': 'PARENT_NETPROFIT', '除税后溢利': 'NETPROFIT',
    '净利润': 'NETPROFIT', '持续经营业务税后利润': 'NETPROFIT', '少数股东损益': 'MINORITY_INTEREST',
    '利息收入': 'INTEREST_INCOME', '融资成本': 'FE_INTEREST_EXPENSE', '利息支出': 'FE_INTEREST_EXPENSE',
    '销售及分销费用': 'SALE_EXPENSE', '销售费用': 'SALE_EXPENSE', '行政开支': 'MANAGE_EXPENSE',
    '管理费用': 'MANAGE_EXPENSE', '研发费用': 'RESEARCH_EXPENSE', '其他收益': 'OTHER_INCOME',
    '应占联营公司溢利': 'INVEST_INCOME', '投资收益': 'INVEST_INCOME',
    '经营业务现金净额': 'NETCASH_OPERATE', '经营活动现金流量净额': 'NETCASH_OPERATE',
    '投资业务现金净额': 'NETCASH_INVEST', '投资活动现金流量净额': 'NETCASH_INVEST',
    '融资业务现金净额': 'NETCASH_FINANCE', '筹资活动现金流量净额': 'NETCASH_FINANCE',
    '期末现金': 'END_CCE', '期末现金及现金等价物': 'END_CCE', '期初现金': 'BEGIN_CCE',
    '现金净额': 'CCE_ADD', '现金及现金等价物净增加额': 'CCE_ADD',
    '购建固定资产': 'CONSTRUCT_LONG_ASSET', '购建固定资产等支付的现金': 'CONSTRUCT_LONG_ASSET',
    '投资支付现金': 'INVEST_PAY_CASH', '收回投资所得现金': 'WITHDRAW_INVEST',
    '新增借款': 'RECEIVE_LOAN_CASH', '取得借款收到的现金': 'RECEIVE_LOAN_CASH',
    '偿还借款': 'PAY_DEBT_CASH', '偿还债务支付的现金': 'PAY_DEBT_CASH',
    '已付股息(融资)': 'ASSIGN_DIVIDEND_PORFIT', '分配股利利润或偿付利息支付的现金': 'ASSIGN_DIVIDEND_PORFIT',
    '发行债券': 'ISSUE_BOND',     '收购附属公司': 'OBTAIN_SUBSIDIARY_OTHER',
    '取得子公司支付的现金净额': 'OBTAIN_SUBSIDIARY_OTHER',
    # 补充：实盘中出现的港交所标准科目名
    '物业厂房及设备': 'FIXED_ASSET', '土地使用权': 'INTANGIBLE_ASSET',
    '合营公司权益': 'LONG_EQUITY_INVEST', '联营公司权益': 'LONG_EQUITY_INVEST',
    '指定以公允价值记账之金融资产': 'TRADABLE_FINASSET',
    '指定以公允价值记账之金融资产(流动)': 'TRADABLE_FINASSET',
    '其他金融资产(流动)': 'TRADABLE_FINASSET', '中长期存款': 'MONETARYFUNDS',
    '其他金融资产(非流动)': 'OTHER_NONCURRENT_ASSET',
    '预付款按金及其他应收款': 'OTHER_RECE',
    '受限制存款及现金': 'MONETARYFUNDS', '现金及等价物': 'MONETARYFUNDS',
    '短期存款': 'MONETARYFUNDS', '递延收入(流动)': 'DEFER_INCOME',
    '递延收入(非流动)': 'DEFER_INCOME', '其他应付款及应计费用': 'OTHER_PAYABLE',
    '其他金融负债(流动)': 'CURRENT_LIAB_OTHER', '融资租赁负债(非流动)': 'LEASE_LIAB',
    '其他营业收入': 'OTHER_INCOME', '应占合营公司溢利': 'INVEST_INCOME',
    '全面收益总额': 'TOTAL_COMPRE_INCOME', '处置固定资产': 'DISPOSAL_LONG_ASSET',
    '出售附属公司': 'DISPOSAL_LONG_ASSET', '已付税项': 'PAY_ALL_TAX',
    '已收股息(投资)': 'RECEIVE_INVEST_INCOME', '购建无形资产及其他资产': 'CONSTRUCT_LONG_ASSET',
}

# 港股「总计/合计」行识别（用于加粗+浅色底）
HK_TOTAL = {'总资产', '资产总计', '总负债', '负债总计', '净资产', '总权益', '权益总额',
            '股东权益', '股东权益合计', '总权益及总负债', '权益及负债合计',
            '负债和所有者权益总计', '负债及权益总计'}


# 标准财报科目顺序（与财报 PDF 一致）：仅展示这些科目，按此顺序排列；其余内部汇总列/空列自动跳过。
CANON_ORDER = {
    'balance': [
        ('MONETARYFUNDS', '货币资金'), ('TRADABLE_FINASSET', '交易性金融资产'),
        ('DERIVE_FINASSET', '衍生金融资产'), ('NOTE_RECE', '应收票据'),
        ('ACCOUNTS_RECE', '应收账款'), ('NOTE_ACCOUNTS_RECE', '应收款项融资'),
        ('PREPAYMENT', '预付款项'), ('OTHER_RECE', '其他应收款'),
        ('DIVIDEND_RECE', '应收股利'), ('INTEREST_RECE', '应收利息'),
        ('CONTRACT_ASSET', '合同资产'), ('HOLDSALE_ASSET', '持有待售资产'),
        ('NONCURRENT_ASSET_1YEAR', '一年内到期的非流动资产'), ('OTHER_CURRENT_ASSET', '其他流动资产'),
        ('TOTAL_CURRENT_ASSETS', '流动资产合计'),
        ('LOAN_ADVANCE', '发放贷款及垫款'), ('AMORTIZE_COST_FINASSET', '债权投资'),
        ('FVTOCI_FINASSET', '其他债权投资'), ('LONG_RECE', '长期应收款'),
        ('LONG_EQUITY_INVEST', '长期股权投资'), ('OTHER_EQUITY_INVEST', '其他权益工具投资'),
        ('FVTPL_FINASSET', '其他非流动金融资产'), ('INVEST_REALESTATE', '投资性房地产'),
        ('FIXED_ASSET', '固定资产'), ('CIP', '在建工程'),
        ('CONSUMPTIVE_BIOLOGICAL_ASSET', '消耗性生物资产'), ('PRODUCTIVE_BIOLOGY_ASSET', '生产性生物资产'),
        ('OIL_GAS_ASSET', '油气资产'), ('USERIGHT_ASSET', '使用权资产'),
        ('INTANGIBLE_ASSET', '无形资产'), ('DEVELOP_EXPENSE', '开发支出'),
        ('GOODWILL', '商誉'), ('LONG_PREPAID_EXPENSE', '长期待摊费用'),
        ('DEFER_TAX_ASSET', '递延所得税资产'), ('OTHER_NONCURRENT_ASSET', '其他非流动资产'),
        ('TOTAL_NONCURRENT_ASSETS', '非流动资产合计'), ('TOTAL_ASSETS', '资产总计'),
        ('SHORT_LOAN', '短期借款'), ('LOAN_PBC', '向中央银行借款'),
        ('BORROW_FUND', '拆入资金'), ('ACCEPT_DEPOSIT_INTERBANK', '吸收存款及同业存放'),
        ('TRADE_FINLIAB', '交易性金融负债'), ('DERIVE_FINLIAB', '衍生金融负债'),
        ('NOTE_PAYABLE', '应付票据'), ('ACCOUNTS_PAYABLE', '应付账款'),
        ('ADVANCE_RECEIVABLES', '预收款项'), ('CONTRACT_LIAB', '合同负债'),
        ('STAFF_SALARY_PAYABLE', '应付职工薪酬'), ('TAX_PAYABLE', '应交税费'),
        ('INTEREST_PAYABLE', '应付利息'), ('DIVIDEND_PAYABLE', '应付股利'),
        ('OTHER_PAYABLE', '其他应付款'), ('HOLDSALE_LIAB', '持有待售负债'),
        ('NONCURRENT_LIAB_1YEAR', '一年内到期的非流动负债'), ('CURRENT_LIAB_OTHER', '其他流动负债'),
        ('TOTAL_CURRENT_LIAB', '流动负债合计'),
        ('LONG_LOAN', '长期借款'), ('BOND_PAYABLE', '应付债券'),
        ('LEASE_LIAB', '租赁负债'), ('LONG_PAYABLE', '长期应付款'),
        ('LONG_STAFFSALARY_PAYABLE', '长期应付职工薪酬'), ('PREDICT_LIAB', '预计负债'),
        ('DEFER_INCOME', '递延收益'), ('DEFER_TAX_LIAB', '递延所得税负债'),
        ('OTHER_NONCURRENT_LIAB', '其他非流动负债'), ('TOTAL_NONCURRENT_LIAB', '非流动负债合计'),
        ('TOTAL_LIABILITIES', '负债合计'),
        ('SHARE_CAPITAL', '股本'), ('OTHER_EQUITY_TOOL', '其他权益工具'),
        ('CAPITAL_RESERVE', '资本公积'), ('TREASURY_SHARES', '库存股'),
        ('OTHER_COMPRE_INCOME', '其他综合收益'), ('SPECIAL_RESERVE', '专项储备'),
        ('SURPLUS_RESERVE', '盈余公积'), ('GENERAL_RISK_RESERVE', '一般风险准备'),
        ('UNASSIGN_RPOFIT', '未分配利润'), ('TOTAL_PARENT_EQUITY', '归属于母公司股东权益合计'),
        ('MINORITY_EQUITY', '少数股东权益'), ('TOTAL_EQUITY', '所有者权益合计'),
        ('TOTAL_LIAB_EQUITY', '负债和所有者权益总计'),
    ],
    'income': [
        ('TOTAL_OPERATE_INCOME', '营业总收入'), ('OPERATE_INCOME', '营业收入'),
        ('TOTAL_OPERATE_COST', '营业总成本'), ('OPERATE_COST', '营业成本'),
        ('OPERATE_TAX_ADD', '税金及附加'), ('SALE_EXPENSE', '销售费用'),
        ('MANAGE_EXPENSE', '管理费用'), ('RESEARCH_EXPENSE', '研发费用'),
        ('FINANCE_EXPENSE', '财务费用'), ('FE_INTEREST_EXPENSE', '利息支出'),
        ('FE_INTEREST_INCOME', '利息收入'), ('ASSET_IMPAIRMENT_LOSS', '资产减值损失'),
        ('CREDIT_IMPAIRMENT_LOSS', '信用减值损失'), ('FAIRVALUE_CHANGE_INCOME', '公允价值变动收益'),
        ('INVEST_INCOME', '投资收益'), ('OTHER_INCOME', '其他收益'),
        ('OPERATE_PROFIT', '营业利润'), ('NONBUSINESS_INCOME', '营业外收入'),
        ('NONBUSINESS_EXPENSE', '营业外支出'), ('TOTAL_PROFIT', '利润总额'),
        ('INCOME_TAX', '所得税费用'), ('NETPROFIT', '净利润'),
        ('PARENT_NETPROFIT', '归母净利润'), ('MINORITY_INTEREST', '少数股东损益'),
        ('DEDUCT_PARENT_NETPROFIT', '扣非归母净利润'), ('BASIC_EPS', '基本每股收益'),
        ('DILUTED_EPS', '稀释每股收益'), ('OTHER_COMPRE_INCOME', '其他综合收益'),
        ('TOTAL_COMPRE_INCOME', '综合收益总额'),
    ],
    'cash': [
        ('SALES_SERVICES', '销售商品提供劳务收到的现金'), ('RECEIVE_TAX_REFUND', '收到的税费返还'),
        ('RECEIVE_OTHER_OPERATE', '收到其他与经营活动有关的现金'), ('TOTAL_OPERATE_INFLOW', '经营活动现金流入小计'),
        ('BUY_SERVICES', '购买商品接受劳务支付的现金'), ('PAY_STAFF_CASH', '支付给职工以及为职工支付的现金'),
        ('PAY_ALL_TAX', '支付的各项税费'), ('PAY_OTHER_OPERATE', '支付其他与经营活动有关的现金'),
        ('TOTAL_OPERATE_OUTFLOW', '经营活动现金流出小计'), ('NETCASH_OPERATE', '经营活动现金流量净额'),
        ('WITHDRAW_INVEST', '收回投资收到的现金'), ('RECEIVE_INVEST_INCOME', '取得投资收益收到的现金'),
        ('DISPOSAL_LONG_ASSET', '处置固定资产等收回的现金净额'), ('RECEIVE_OTHER_INVEST', '收到其他与投资活动有关的现金'),
        ('TOTAL_INVEST_INFLOW', '投资活动现金流入小计'), ('CONSTRUCT_LONG_ASSET', '购建固定资产等支付的现金'),
        ('INVEST_PAY_CASH', '投资支付的现金'), ('OBTAIN_SUBSIDIARY_OTHER', '取得子公司支付的现金净额'),
        ('PAY_OTHER_INVEST', '支付其他与投资活动有关的现金'), ('TOTAL_INVEST_OUTFLOW', '投资活动现金流出小计'),
        ('NETCASH_INVEST', '投资活动现金流量净额'),
        ('RECEIVE_LOAN_CASH', '取得借款收到的现金'), ('ISSUE_BOND', '发行债券收到的现金'),
        ('RECEIVE_OTHER_FINANCE', '收到其他与筹资活动有关的现金'), ('TOTAL_FINANCE_INFLOW', '筹资活动现金流入小计'),
        ('PAY_DEBT_CASH', '偿还债务支付的现金'), ('ASSIGN_DIVIDEND_PORFIT', '分配股利利润或偿付利息支付的现金'),
        ('PAY_OTHER_FINANCE', '支付其他与筹资活动有关的现金'), ('TOTAL_FINANCE_OUTFLOW', '筹资活动现金流出小计'),
        ('NETCASH_FINANCE', '筹资活动现金流量净额'),
        ('CCE_ADD', '现金及现金等价物净增加额'), ('BEGIN_CCE', '期初现金及现金等价物余额'),
        ('END_CCE', '期末现金及现金等价物余额'),
    ],
}


# 总计/合计行识别：用于前端与 Excel 加粗 + 浅色背景
TOTAL_SUFFIX = ('合计', '总计')

# 资产负债表「主要合计」在最前面重复列出的科目（按此顺序）
BALANCE_TOP = [
    '资产总计', '负债合计', '归属于母公司股东权益合计', '流动资产合计',
    '非流动资产合计', '流动负债合计', '非流动负债合计', '固定资产',
    '负债和所有者权益总计',
]

# 利润表「主要合计/小计」在最前面重复列出的科目（按此顺序）
INCOME_TOP = [
    '营业总收入', '营业总成本', '营业利润', '利润总额', '净利润',
    '归母净利润', '扣非归母净利润', '综合收益总额',
]

# 现金流量表「主要合计/小计」在最前面重复列出的科目（按此顺序）
CASH_TOP = [
    '经营活动现金流入小计', '经营活动现金流出小计', '经营活动现金流量净额',
    '投资活动现金流入小计', '投资活动现金流出小计', '投资活动现金流量净额',
    '筹资活动现金流入小计', '筹资活动现金流出小计', '筹资活动现金流量净额',
    '现金及现金等价物净增加额', '期初现金及现金等价物余额', '期末现金及现金等价物余额',
]

# 三表顶部摘要科目汇总（用于：前端顶部摘要卡片 + Excel 顶部重复行 + 收入/现金流表的加粗识别）
SUMMARY_TOP = {'资产负债表': BALANCE_TOP, '利润表': INCOME_TOP, '现金流量表': CASH_TOP}
# 收入表 / 现金流量表的关键「总计、小计」行集合（用于加粗 + 浅色背景，区别于资产负债表的后缀匹配）
INCOME_CASH_SUMMARY = set(INCOME_TOP) | set(CASH_TOP)


def _norm_digits(s):
    s = (s or '').strip()
    m = __import__('re').search(r'\d{6}', s)
    return m.group(0) if m else s


def _resolve_company_input(raw, cache=None):
    """把用户输入解析成 6 位 A 股代码，同时返回标准中文名称。
    支持：6 位代码、带前缀/后缀的代码（如 SH601899 / 601899.SZ / 紫金矿业(SH601899)）、
    以及纯中文名称（依赖 peers_cache.json 的 name_to_code 索引，或在线回退）。
    """
    raw = (raw or '').strip()
    if not raw:
        raise RuntimeError('请输入公司代码或名称。')
    # 情况 1：已经能提取 6 位数字
    digits = _norm_digits(raw)
    if digits != raw and re.fullmatch(r'\d{6}', digits):
        return _resolve_with_name(digits, cache)
    # 情况 2：纯 6 位数字
    if re.fullmatch(r'\d{6}', raw):
        return _resolve_with_name(raw, cache)
    # 情况 3：中文名称（可能在缓存 name_to_code 中）
    if cache is None:
        cache = _load_peers_cache()
    name_to_code = cache.get('name_to_code')
    if name_to_code is None:
        # 旧版缓存没有反向索引，动态生成一次
        name_to_code = {name.strip(): code for code, name in cache.get('code_to_name', {}).items()}
        cache['name_to_code'] = name_to_code
    code = name_to_code.get(raw)
    if code:
        return code, cache.get('code_to_name', {}).get(code, raw)
    # 情况 4：在线回退（东方财富数据中心，与财报接口同源，通常可达）
    try:
        df = ak.stock_info_a_code_name()
        mask = df['name'].astype(str).str.strip() == raw
        if mask.any():
            row = df[mask].iloc[0]
            return str(row['code']).strip(), str(row['name']).strip()
    except Exception:
        pass
    raise RuntimeError(f'未找到「{raw}」对应的 A 股代码。请确认名称正确，或改用 6 位代码。')


def _resolve_with_name(code, cache):
    if cache is None:
        cache = _load_peers_cache()
    name = cache.get('code_to_name', {}).get(code)
    if not name:
        # 在线回退查名称
        try:
            df = ak.stock_info_a_code_name()
            row = df[df['code'].astype(str).str.strip() == code]
            if not row.empty:
                name = str(row.iloc[0]['name']).strip()
        except Exception:
            pass
    return code, name or code




def _stmt_sym(code):
    code = _norm_digits(code)
    if code[:2] in ('SH', 'SZ', 'BJ'):
        return code
    if code.startswith('6'):
        return 'SH' + code
    if code.startswith(('0', '3')):
        return 'SZ' + code
    if code.startswith(('8', '4')):
        return 'BJ' + code
    return 'SH' + code


# ---------- 港股支持 ----------
def _is_hk(raw):
    """判断输入是否为港股：以 hk 前缀，或纯 5 位数字（A股为 6 位，故 5 位唯一对应港股）。"""
    s = (raw or '').strip().upper()
    if s.startswith('HK'):
        return True
    return bool(re.fullmatch(r'\d{5}', s))


def _norm_hk(raw):
    """港股代码归一化为 5 位（去前缀/补零），如 00700。"""
    s = (raw or '').strip().upper()
    if s.startswith('HK'):
        s = s[2:]
    s = re.sub(r'\D', '', s)
    return s.zfill(5)[-5:] if s else s


_HK_CACHE = None


def _load_hk_cache():
    """加载港股名称缓存（懒加载）。由 hk_stock_cache.json 提供 code<->name。"""
    global _HK_CACHE
    if _HK_CACHE is not None:
        return _HK_CACHE
    here = os.path.dirname(os.path.abspath(__file__))
    p = os.path.join(here, 'hk_stock_cache.json')
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            _HK_CACHE = json.load(f)
    else:
        _HK_CACHE = {}
    return _HK_CACHE


def _resolve_hk_name(stock, hk_cache):
    cache = hk_cache or _load_hk_cache()
    name = cache.get('code_to_name', {}).get(stock)
    if name:
        return name
    # 在线回退：从港股通列表取名称
    try:
        df = ak.stock_hk_famous_spot_em()
        m = df[df['代码'].astype(str).str.zfill(5) == stock]
        if not m.empty:
            return str(m.iloc[0]['名称']).strip()
    except Exception:
        pass
    return stock


def _hk_symbol(kind):
    return {'balance': '资产负债表', 'income': '利润表', 'cash': '现金流量表'}[kind]


def _fetch_hk_statement(stock, kind):
    """拉取港股三表（港交所口径，数值已为港币）。长表 pivot 成 {year:{alias_code:val}}。
    仅保留能映射到 A 股同义代码的科目，从而复用 CANON_ORDER 行序 + _indicator_defs 全部固定指标。

    注：AkShare 港股接口 FISCAL_YEAR 列常为报告期月日字符串（如 '12-31'）而非年份，
    故财年取自 REPORT_DATE 的年份（年报的报告日即对应财年结束时点）。"""
    sym = _hk_symbol(kind)
    df = ak.stock_financial_hk_report_em(stock=stock, symbol=sym, indicator='年度')
    df = df.dropna(subset=['STD_ITEM_NAME', 'REPORT_DATE', 'AMOUNT'])
    df = df.copy()
    df['_YEAR'] = pd.to_datetime(df['REPORT_DATE'], errors='coerce').dt.year
    df = df.dropna(subset=['_YEAR'])
    # 年报去重：同一 (财年, 科目) 取报告日最新一条（避免中期/年度重复）
    df = df.sort_values('REPORT_DATE')
    df = df.drop_duplicates(subset=['_YEAR', 'STD_ITEM_NAME'], keep='last')
    data = {}
    for _, row in df.iterrows():
        y = int(row['_YEAR'])
        name = str(row['STD_ITEM_NAME']).strip()
        alias = HK_ALIAS.get(name)
        if not alias:
            continue
        try:
            v = float(row['AMOUNT'])
        except Exception:
            continue
        data.setdefault(y, {})[alias] = v
    return data


def _fetch_hk_cached(stock, kind, force=False):
    """港股三表：复用本地 SQLite 缓存（code 前缀 HK）。"""
    code = 'HK' + stock
    now = datetime.now()
    if not force and not _cache_stale(code, kind, now):
        cached = _cache_get(code, kind)
        if cached:
            return cached, [item for item, _ in CANON_ORDER.get(kind, [])]
    data = _fetch_hk_statement(stock, kind)
    latest_year = max(data.keys()) if data else 0
    _cache_put(code, kind, data, latest_year)
    return data, [item for item, _ in CANON_ORDER.get(kind, [])]


def _conn():
    """本地 SQLite 缓存连接（惰性创建表）。库文件与 data.py 同目录。"""
    here = os.path.dirname(os.path.abspath(__file__))
    db = os.path.join(here, 'fin_cache.db')
    conn = sqlite3.connect(db)
    conn.execute("""CREATE TABLE IF NOT EXISTS stmt(
        code TEXT, kind TEXT, year INTEGER, item TEXT, value REAL,
        PRIMARY KEY(code,kind,year,item))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS meta(
        code TEXT, kind TEXT, updated_at TEXT, latest_year INTEGER,
        PRIMARY KEY(code,kind))""")
    return conn


def _cache_get(code, kind):
    """读取本地缓存的全部年报：返回 {year:{item:val}}，无则 None。"""
    conn = _conn()
    try:
        rows = conn.execute("SELECT year,item,value FROM stmt WHERE code=? AND kind=?",
                            (code, kind)).fetchall()
    finally:
        conn.close()
    if not rows:
        return None
    data = {}
    for y, item, val in rows:
        data.setdefault(int(y), {})[item] = val
    return data


def _cache_put(code, kind, data, latest_year):
    """写入/覆盖某公司的某张表年报到本地缓存。"""
    conn = _conn()
    try:
        conn.execute("DELETE FROM stmt WHERE code=? AND kind=?", (code, kind))
        for y, items in data.items():
            for item, val in items.items():
                conn.execute("INSERT OR REPLACE INTO stmt VALUES(?,?,?,?,?)",
                             (code, kind, int(y), item, val))
        conn.execute("INSERT OR REPLACE INTO meta VALUES(?,?,?,?)",
                     (code, kind, datetime.now().isoformat(), int(latest_year) if latest_year else 0))
        conn.commit()
    finally:
        conn.close()


def _cache_stale(code, kind, now):
    """判断缓存是否过期。
    年报通常在次年 4 月底前披露：
      - 现在已过 今年5月  → 期望缓存至少含「去年」年报，否则需要补拉；
      - 现在在 5月前      → 期望缓存至少含「前年」年报。
    兜底：缓存时间超过 18 个月也视为过期。
    """
    conn = _conn()
    try:
        r = conn.execute("SELECT updated_at,latest_year FROM meta WHERE code=? AND kind=?",
                         (code, kind)).fetchone()
    finally:
        conn.close()
    if not r:
        return True
    updated = datetime.fromisoformat(r[0]); latest_year = r[1] or 0
    expected = (now.year - 1) if now.month >= 5 else (now.year - 2)
    if latest_year < expected:
        return True
    if (now - updated).days > 540:
        return True
    return False


def _fetch_from_akshare(sym, kind):
    """从 AkShare 联网拉取三表年报，返回 (data, order)。"""
    func = {
        'balance': ak.stock_balance_sheet_by_report_em,
        'income': ak.stock_profit_sheet_by_report_em,
        'cash': ak.stock_cash_flow_sheet_by_report_em,
    }[kind]
    df = func(symbol=sym)
    df = df.copy()
    df['_RD'] = pd.to_datetime(df['REPORT_DATE'], errors='coerce')
    df['_YEAR'] = df['_RD'].dt.year
    ann = df[df['_RD'].dt.month == 12].copy()
    ann = ann.sort_values('_RD')
    item_cols = [c for c in df.columns if c not in METADATA_COLS and not str(c).endswith('_YOY')]
    data = {}
    for _, row in ann.iterrows():
        y = int(row['_YEAR'])
        d = {}
        for c in item_cols:
            v = row[c]
            if pd.isna(v):
                continue
            try:
                fv = float(v)
            except Exception:
                continue
            d[c] = fv
        data[y] = d
    return data, item_cols


def _fetch_statement(sym, kind, force=False):
    """带本地 SQLite 缓存的拉取：首次/过期(或 force)时联网，否则读本地缓存秒开。
    返回 (data, order)，order 仅保留作兼容（展示统一用 CANON_ORDER）。"""
    now = datetime.now()
    if not force and not _cache_stale(sym, kind, now):
        cached = _cache_get(sym, kind)
        if cached:
            return cached, [item for item, _ in CANON_ORDER.get(kind, [])]
    data, _ = _fetch_from_akshare(sym, kind)
    latest_year = max(data.keys()) if data else 0
    _cache_put(sym, kind, data, latest_year)
    return data, [item for item, _ in CANON_ORDER.get(kind, [])]


def _yoy_from_values(vals):
    """vals={year:val} -> {year:pct or None}"""
    years = sorted(vals.keys())
    out = {}
    for i, y in enumerate(years):
        if i == 0 or vals[y] is None:
            out[y] = None
            continue
        prev = vals[years[i - 1]]
        if prev in (None, 0):
            out[y] = None
        else:
            out[y] = (vals[y] - prev) / prev * 100
    return out


def _g(d, k):
    v = d.get(k)
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b


# 固定指标定义：(名称, 单位类型, 计算函数)  计算函数接收 (ib, bb, cb) 三个年度字典
def _indicator_defs():
    def gross(ib, bb, cb):
        rev, cost = _g(ib, 'OPERATE_INCOME'), _g(ib, 'OPERATE_COST')
        return None if (rev in (None, 0)) else (rev - cost) / rev * 100

    def netp(ib, bb, cb):
        rev, np_ = _g(ib, 'OPERATE_INCOME'), _g(ib, 'PARENT_NETPROFIT') or _g(ib, 'NETPROFIT')
        return None if (rev in (None, 0)) else np_ / rev * 100

    def opr(ib, bb, cb):
        rev, op = _g(ib, 'OPERATE_INCOME'), _g(ib, 'OPERATE_PROFIT')
        return None if (rev in (None, 0)) else op / rev * 100

    def roe(ib, bb, cb):
        np_ = _g(ib, 'PARENT_NETPROFIT') or _g(ib, 'NETPROFIT')
        eq = _g(bb, 'TOTAL_PARENT_EQUITY') or _g(bb, 'TOTAL_EQUITY')
        r = _div(np_, eq)
        return r * 100 if r is not None else None

    def roa(ib, bb, cb):
        np_ = _g(ib, 'PARENT_NETPROFIT') or _g(ib, 'NETPROFIT')
        ta = _g(bb, 'TOTAL_ASSETS')
        r = _div(np_, ta)
        return r * 100 if r is not None else None

    def dupont_np(ib, bb, cb):
        return netp(ib, bb, cb)

    def dupont_turn(ib, bb, cb):
        rev, ta = _g(ib, 'OPERATE_INCOME'), _g(bb, 'TOTAL_ASSETS')
        return _div(rev, ta)

    def dupont_eqm(ib, bb, cb):
        ta = _g(bb, 'TOTAL_ASSETS')
        eq = _g(bb, 'TOTAL_PARENT_EQUITY') or _g(bb, 'TOTAL_EQUITY')
        return _div(ta, eq)

    def debt_ratio(ib, bb, cb):
        tl, ta = _g(bb, 'TOTAL_LIABILITIES'), _g(bb, 'TOTAL_ASSETS')
        r = _div(tl, ta)
        return r * 100 if r is not None else None

    def current_ratio(ib, bb, cb):
        ca, cl = _g(bb, 'TOTAL_CURRENT_ASSETS'), _g(bb, 'TOTAL_CURRENT_LIAB')
        return _div(ca, cl)

    def quick_ratio(ib, bb, cb):
        ca, inv, cl = _g(bb, 'TOTAL_CURRENT_ASSETS'), _g(bb, 'INVENTORY'), _g(bb, 'TOTAL_CURRENT_LIAB')
        if None in (ca, cl) or cl == 0:
            return None
        inv = inv or 0
        return (ca - inv) / cl

    def interest_cover(ib, bb, cb):
        tp, ie = _g(ib, 'TOTAL_PROFIT'), _g(ib, 'FE_INTEREST_EXPENSE')
        if None in (tp, ie) or ie == 0:
            return None
        return (tp + ie) / ie

    def ar_turn(ib, bb, cb):
        rev, ar = _g(ib, 'OPERATE_INCOME'), _g(bb, 'ACCOUNTS_RECE')
        return _div(rev, ar)

    def inv_turn(ib, bb, cb):
        cost, inv = _g(ib, 'OPERATE_COST'), _g(bb, 'INVENTORY')
        return _div(cost, inv)

    def ap_turn(ib, bb, cb):
        cost, ap = _g(ib, 'OPERATE_COST'), _g(bb, 'ACCOUNTS_PAYABLE')
        return _div(cost, ap)

    def ar_days(ib, bb, cb):
        t = ar_turn(ib, bb, cb)
        return None if t in (None, 0) else 365 / t

    def inv_days(ib, bb, cb):
        t = inv_turn(ib, bb, cb)
        return None if t in (None, 0) else 365 / t

    def ap_days(ib, bb, cb):
        t = ap_turn(ib, bb, cb)
        return None if t in (None, 0) else 365 / t

    def ccc(ib, bb, cb):
        a, i, p = ar_days(ib, bb, cb), inv_days(ib, bb, cb), ap_days(ib, bb, cb)
        return None if None in (a, i, p) else a + i - p

    def ar_ap_diff(ib, bb, cb):
        # 应收合计：所有"应收/预付"类资产（对手方欠公司的钱 + 公司预付/垫款）
        rece_keys = ('NOTE_RECE', 'ACCOUNTS_RECE', 'NOTE_ACCOUNTS_RECE', 'PREPAYMENT',
                     'OTHER_RECE', 'DIVIDEND_RECE', 'INTEREST_RECE', 'CONTRACT_ASSET', 'LONG_RECE')
        # 应付合计：所有"应付/预收"类负债（公司欠对手方的钱 + 客户预付/合同负债 + 各项计提）
        pay_keys = ('NOTE_PAYABLE', 'ACCOUNTS_PAYABLE', 'ADVANCE_RECEIVABLES', 'CONTRACT_LIAB',
                    'STAFF_SALARY_PAYABLE', 'TAX_PAYABLE', 'INTEREST_PAYABLE', 'DIVIDEND_PAYABLE',
                    'OTHER_PAYABLE', 'LONG_PAYABLE')
        rece = sum((_g(bb, k) or 0) for k in rece_keys)
        pay = sum((_g(bb, k) or 0) for k in pay_keys)
        if rece == 0 and pay == 0:
            return None
        # 正数：被上下游净占用的营运资金（自有现金被垫住）；负数：更占用供应商/客户资金（对现金更友好）
        return rece - pay

    def rev_abs(ib, bb, cb):
        return _g(ib, 'OPERATE_INCOME')

    def np_abs(ib, bb, cb):
        return _g(ib, 'PARENT_NETPROFIT') or _g(ib, 'NETPROFIT')

    def ta_abs(ib, bb, cb):
        return _g(bb, 'TOTAL_ASSETS')

    def ocf_np(ib, bb, cb):
        ocf = _g(cb, 'NETCASH_OPERATE')
        np_ = _g(ib, 'PARENT_NETPROFIT') or _g(ib, 'NETPROFIT')
        return _div(ocf, np_)

    def fcf(ib, bb, cb):
        ocf = _g(cb, 'NETCASH_OPERATE')
        capex = _g(cb, 'CONSTRUCT_LONG_ASSET')
        return None if (ocf is None or capex is None) else ocf - capex

    def capex_ratio(ib, bb, cb):
        capex, rev = _g(cb, 'CONSTRUCT_LONG_ASSET'), _g(ib, 'OPERATE_INCOME')
        return None if (rev in (None, 0)) else capex / rev * 100

    groups = [
        ('盈利质量', [
            ('毛利率(%)', gross, '毛利率 = (营业收入 − 营业成本) ÷ 营业收入 × 100%。反映产品本身的赚钱空间，越高越好。'),
            ('净利率(%)', netp, '净利率 = 归母净利润 ÷ 营业收入 × 100%。扣除全部费用与税后的真实利润率。'),
            ('营业利润率(%)', opr, '营业利润率 = 营业利润 ÷ 营业收入 × 100%。仅看主营业务的盈利（不含投资收益、营业外收支）。'),
            ('ROE(%)', roe, '净资产收益率 = 归母净利润 ÷ 归属于母公司股东权益 × 100%。股东每投 1 元能赚多少，是核心盈利能力指标。'),
            ('ROA(%)', roa, '总资产收益率 = 归母净利润 ÷ 总资产 × 100%。衡量全部资产（含借来的钱）的整体回报。'),
        ]),
        ('杜邦拆解', [
            ('净利率(%)', dupont_np, '同「净利率」，杜邦分析把它作为驱动 ROE 的第一个因子（ROE = 净利率 × 总资产周转率 × 权益乘数）。'),
            ('总资产周转率(次)', dupont_turn, '总资产周转率 = 营业收入 ÷ 期末总资产。资产一年能"转"出几倍收入，越高越会用资产。注：严格杜邦用期初与期末平均总资产，本工具为简化用期末总资产。'),
            ('权益乘数(倍)', dupont_eqm, '权益乘数 = 总资产 ÷ 归属于母公司股东权益。反映加杠杆倍数，越高=借的钱越多、财务风险越大。'),
        ]),
        ('偿债/资本结构', [
            ('资产负债率(%)', debt_ratio, '资产负债率 = 总负债 ÷ 总资产 × 100%。越高=负债占比越大、长期偿债压力越大。'),
            ('流动比率(倍)', current_ratio, '流动比率 = 流动资产 ÷ 流动负债。衡量 1 年内能否还上短期债，通常 >1 较安全。'),
            ('速动比率(倍)', quick_ratio, '速动比率 = (流动资产 − 存货) ÷ 流动负债。比流动比率更严，剔除变现慢的存货。'),
            ('利息保障倍数(倍)', interest_cover, '利息保障倍数 = (利润总额 + 利息费用) ÷ 利息费用。赚的钱够付几倍利息，<1 说明还息有风险。'),
        ]),
        ('运营效率', [
            ('应收账款周转率(次)', ar_turn, '应收账款周转率 = 营业收入 ÷ 应收账款。一年把赊账收回再借出的次数，越高=回款越快。'),
            ('存货周转率(次)', inv_turn, '存货周转率 = 营业成本 ÷ 存货。一年存货卖空几次，越高=积压越少。'),
            ('应付账款周转天数(天)', ap_days, '应付天数 = 365 ÷ (营业成本 ÷ 应付账款)。平均花多少天付供应商货款，越长=越占用供应商资金。'),
            ('现金循环周期(天)', ccc, '现金循环周期 = 应收天数 + 存货天数 − 应付天数。从付钱买料到收回货款的净现金占用天数，越短越省现金。'),
            ('应收和应付的差额(元)', ar_ap_diff, '应收-应付差额 = 应收合计 − 应付合计（绝对值，单位与财报一致）。应收合计 = 应收票据+应收账款+应收款项融资+预付款项+其他应收款+应收股利+应收利息+合同资产+长期应收款；应付合计 = 应付票据+应付账款+预收款项+合同负债+应付职工薪酬+应交税费+应付利息+应付股利+其他应付款+长期应付款。正数=被上下游净占用的营运资金（自有现金被垫住）；负数=更占用供应商/客户资金（对现金更友好）。'),
        ]),
        ('成长能力', [
            ('营业收入(元)', rev_abs, '利润表"营业总收入"（绝对值，单位元）。规模本身，增长看同比(%)折线。'),
            ('归母净利润(元)', np_abs, '利润表"归属于母公司股东的净利润"（绝对值）。'),
            ('总资产(元)', ta_abs, '资产负债表"资产总计"（绝对值）。'),
        ]),
        ('现金流质量', [
            ('经营现金流/净利润(倍)', ocf_np, '经营现金流净额 ÷ 归母净利润。>1 说明赚的是真金白银（利润含金量高），长期 <1 要警惕利润虚。'),
            ('自由现金流(元)', fcf, '自由现金流 = 经营活动现金流量净额 − 购建长期资产支付的现金。企业真正能自由支配的现金。'),
            ('资本支出占比(%)', capex_ratio, '资本支出占比 = 购建长期资产支付的现金 ÷ 营业收入 × 100%。扩产/维护投入占收入比重。'),
        ]),
    ]
    return groups


def get_comparison(companies, years):
    companies = [c for c in companies if c]
    raw = {}
    comp_list = []
    cache = _load_peers_cache()
    hk_cache = _load_hk_cache()
    for c in companies:
        raw_input = (c.get('code') or c.get('name') or '').strip()
        # 港股分支：5 位代码 / hk 前缀
        if _is_hk(raw_input):
            stock = _norm_hk(raw_input)
            name = _resolve_hk_name(stock, hk_cache)
            # 仅当用户传入的是真实名称（非代码本身）才采用，避免把代码误当名称
            nm = str(c.get('name') or '').strip()
            if nm and not _is_hk(nm):
                name = nm
            comp_list.append({'code': stock, 'name': name, 'market': '港股', 'currency': 'HKD'})
            raw[stock] = {
                'balance': _fetch_hk_cached(stock, 'balance'),
                'income': _fetch_hk_cached(stock, 'income'),
                'cash': _fetch_hk_cached(stock, 'cash'),
            }
            continue
        # A 股分支
        code, name = _resolve_company_input(raw_input, cache)
        # 若用户显式传了名称，以用户为准
        if c.get('name') and str(c.get('name')).strip():
            name = str(c.get('name')).strip()
        comp_list.append({'code': code, 'name': name, 'market': 'A股', 'currency': 'CNY'})
        sym = _stmt_sym(code)
        raw[code] = {
            'balance': _fetch_statement(sym, 'balance'),
            'income': _fetch_statement(sym, 'income'),
            'cash': _fetch_statement(sym, 'cash'),
        }

    # 统一的年份列表：所有公司年报年份的并集，取最近 years 个
    all_years = set()
    for code, kinds in raw.items():
        for k in kinds:
            all_years.update(kinds[k][0].keys())
    years_list = sorted(all_years)[-max(1, int(years)):]

    EXCLUDE_SUFFIX = ('_BALANCE', '_NOTE', '_OTHER')

    def _collect(kind, item):
        vals, rawvals, keep = {}, {}, False
        for comp in comp_list:
            d_year = raw[comp['code']][kind][0]  # {year: {item: val}}
            per_year = {y: d_year.get(y, {}).get(item) for y in years_list}
            rawvals[comp['name']] = per_year
            vals[comp['name']] = per_year
            for y in years_list:
                v = per_year[y]
                if v not in (None, 0):
                    keep = True
        return vals, rawvals, keep

    def build_statement_rows(kind):
        rows = []
        shown = set()
        # 1) 按标准财报顺序展示已知科目
        for item, name in CANON_ORDER.get(kind, []):
            vals, rawvals, keep = _collect(kind, item)
            if not keep:
                continue
            yoy = {c['name']: _yoy_from_values(rawvals[c['name']]) for c in comp_list}
            rows.append({'name': name, 'values': vals, 'yoy': yoy,
                         'is_total': (name.endswith(TOTAL_SUFFIX)
                                      or (kind != 'balance' and name in INCOME_CASH_SUMMARY))})
            shown.add(item)
        # 2) 兜底：财报标准清单未覆盖、但确有数据且已命名（中文）的科目，按代码排序追加
        all_fields = set()
        for comp in comp_list:
            all_fields.update(raw[comp['code']][kind][0].keys())
        for item in sorted(all_fields):
            if item in shown or item not in NAME_MAP:
                continue
            if item.endswith(EXCLUDE_SUFFIX) or item.startswith('TOTAL_OTHER_'):
                continue
            vals, rawvals, keep = _collect(kind, item)
            if not keep:
                continue
            yoy = {c['name']: _yoy_from_values(rawvals[c['name']]) for c in comp_list}
            rows.append({'name': NAME_MAP[item], 'values': vals, 'yoy': yoy,
                         'is_total': (NAME_MAP[item].endswith(TOTAL_SUFFIX)
                                      or (kind != 'balance' and NAME_MAP[item] in INCOME_CASH_SUMMARY))})
        return rows

    tables = {}
    for kind, label in [('balance', '资产负债表'), ('income', '利润表'), ('cash', '现金流量表')]:
        tables[label] = build_statement_rows(kind)

    # 固定指标 + 杜邦（每个指标一行，含所有公司）
    groups = _indicator_defs()
    fixed_rows = []
    dupont_rows = []
    src = {}
    for comp in comp_list:
        code = comp['code']
        ib = raw[code]['income'][0]
        bb = raw[code]['balance'][0]
        cb = raw[code]['cash'][0]
        src[comp['name']] = {y: (ib.get(y, {}), bb.get(y, {}), cb.get(y, {})) for y in years_list}
    for gname, defs in groups:
        target = fixed_rows if gname != '杜邦拆解' else dupont_rows
        for rname, fn, note in defs:
            vals = {}
            yoyv = {}
            for comp in comp_list:
                per_year = {y: fn(*src[comp['name']][y]) for y in years_list}
                vals[comp['name']] = per_year
                yoyv[comp['name']] = _yoy_from_values(per_year)
            target.append({'name': rname, 'values': vals, 'yoy': yoyv, 'note': note, 'group': gname})
    tables['固定指标'] = fixed_rows
    tables['杜邦分析'] = dupont_rows

    return {'companies': comp_list, 'years': years_list, 'tables': tables,
            'summary_top': SUMMARY_TOP}


_PEERS_CACHE = None


def _load_peers_cache():
    """加载离线同行缓存（懒加载，单次进程内只加载一次）。
    同时保证缓存里包含 name_to_code（名称 -> 代码）反向索引，兼容旧版缓存。"""
    global _PEERS_CACHE
    if _PEERS_CACHE is not None:
        return _PEERS_CACHE
    here = os.path.dirname(os.path.abspath(__file__))
    p = os.path.join(here, 'peers_cache.json')
    if not os.path.exists(p):
        raise RuntimeError('同行缓存文件缺失（peers_cache.json）。请先运行 build_peers_cache.py 生成缓存，'
                           '或改用「手动加公司」选择可比公司。')
    with open(p, 'r', encoding='utf-8') as f:
        _PEERS_CACHE = json.load(f)
    # 兼容旧缓存：动态补 name_to_code
    if 'name_to_code' not in _PEERS_CACHE:
        _PEERS_CACHE['name_to_code'] = {
            name.strip(): code for code, name in _PEERS_CACHE.get('code_to_name', {}).items()
        }
    return _PEERS_CACHE


def get_peers(raw):
    """根据股票代码或名称返回所属申万二级行业及同行公司（全部来自离线缓存，运行时零网络）。"""
    cache = _load_peers_cache()
    code, primary_name = _resolve_company_input(raw, cache)
    meta = cache.get('code_to_industry', {}).get(code)
    if not meta:
        raise RuntimeError(f'代码 {code} 不在申万行业缓存中（可能是新股 / 北交所 / 港股美股）。'
                           f'请改用「手动加公司」选择可比公司。')
    ic = meta['industry_code']
    industry_name = meta.get('industry_name') or ic
    members = cache.get('industry_to_peers', {}).get(ic, [])
    peers = [{'code': m['code'], 'name': m['name']} for m in members if m['code'] != code]
    return {'industry': industry_name, 'primaryName': primary_name, 'peers': peers[:60]}


def build_excel(body):
    """body 结构同 get_comparison 返回；每个表写 数值 + 同比 两个透视块。
    注意：经 HTTP/JSON 往返后，values/yoy 里的年份键会从 int 变成 str，
    这里统一归一化为 int，避免 .get(y) 用 int 去匹配 str 键而全部落空。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    companies = body['companies']
    years = body['years']

    def _norm_key(k):
        try:
            return int(k)
        except Exception:
            return k

    # 归一化所有年份键为 int（兼容 JSON 往返后的 str 键）
    for rows in body.get('tables', {}).values():
        for r in rows:
            for comp in companies:
                nm = comp.get('name')
                if nm and r.get('values') and nm in r['values']:
                    r['values'][nm] = {_norm_key(y): v for y, v in r['values'][nm].items()}
                if nm and r.get('yoy') and nm in r['yoy']:
                    r['yoy'][nm] = {_norm_key(y): v for y, v in r['yoy'][nm].items()}

    head_fill = PatternFill('solid', fgColor='E6F1FB')
    sub_fill = PatternFill('solid', fgColor='EEF1F4')
    name_fill = PatternFill('solid', fgColor='F7F8FA')
    total_fill = PatternFill('solid', fgColor='DCE9F7')
    bold_font = Font(bold=True)
    thin = Side(style='thin', color='D3D1C7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 与网页图表图例一致的调色板（主公司=首个）
    palette = ['185FA5', '0F6E56', 'B45309', '7F77DD', '993556']
    def _lighten(hex6, ratio):
        r = int(hex6[0:2], 16); g = int(hex6[2:4], 16); b = int(hex6[4:6], 16)
        r = int(r + (255 - r) * ratio); g = int(g + (255 - g) * ratio); b = int(b + (255 - b) * ratio)
        return '%02X%02X%02X' % (r, g, b)
    # 各公司数据列的浅色底（主公司略深，便于区分）
    comp_fill = [PatternFill('solid', fgColor=_lighten(p, 0.82 if i == 0 else 0.90))
                 for i, p in enumerate(palette)]
    white_bold = Font(bold=True, color='FFFFFF')

    def write_sheet(ws, title, rows, show_yoy):
        ws.append([f'{title} · ' + ('同比(%)' if show_yoy else '数值')])
        ws.cell(1, 1).font = Font(bold=True, size=13, color='0C447C')
        # header row 2: year groups（每个年份跨公司子列合并居中）
        h1 = [title + ' \\ 年份']
        for _y in years:
            h1.append(_y)
            h1.extend(['' for _ in range(len(companies) - 1)])
        ws.append(h1)
        # header row 3: companies
        h2 = [''] + [c['name'] for _y in years for c in companies]
        ws.append(h2)
        # 合并年份分组单元格并居中
        col = 2
        for _y in years:
            if len(companies) > 1:
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + len(companies) - 1)
            ws.cell(2, col).alignment = Alignment(horizontal='center')
            col += len(companies)
        for ci in range(1, len(h2) + 1):
            ws.cell(2, ci).fill = head_fill
            ws.cell(3, ci).fill = sub_fill
            ws.cell(2, ci).border = border
            ws.cell(3, ci).border = border
        ws.cell(2, 1).fill = head_fill
        ws.cell(3, 1).fill = sub_fill
        # 公司列表头（第3行）：按调色板实心色 + 白字，主公司加粗
        for y_idx in range(len(years)):
            for cidx, c in enumerate(companies):
                col = 2 + y_idx * len(companies) + cidx
                cell = ws.cell(3, col)
                cell.fill = PatternFill('solid', fgColor=palette[cidx % len(palette)])
                cell.font = white_bold if cidx == 0 else Font(color='FFFFFF')
        cols_per_year = len(companies)
        for r in rows:
            src = r['yoy'] if show_yoy else r['values']
            row = [r['name']]
            for y in years:
                for c in companies:
                    v = src.get(c['name'], {}).get(y)
                    row.append(None if v is None else (round(v, 2) if isinstance(v, float) else v))
            ws.append(row)
            rr = ws.max_row
            is_total = r.get('is_total')
            # 总计/小计行：科目名单元格也加粗 + 浅蓝底（与网页 .total-row 整行高亮一致）
            ws.cell(rr, 1).fill = total_fill if is_total else name_fill
            if is_total:
                ws.cell(rr, 1).font = bold_font
            ws.cell(rr, 1).border = border
            for y_idx in range(len(years)):
                for cidx, c in enumerate(companies):
                    col = 2 + y_idx * cols_per_year + cidx
                    cell = ws.cell(rr, col)
                    cell.border = border
                    cell.number_format = '#,##0.00'
                    if is_total:
                        # 总计/合计行：加粗 + 统一浅蓝（主公司仍加粗）
                        cell.font = bold_font
                        cell.fill = total_fill
                    else:
                        # 普通行：按公司列上浅色底；主公司数字加粗
                        cell.fill = comp_fill[cidx % len(companies)]
                        if cidx == 0:
                            cell.font = bold_font

    for tname, rows in body['tables'].items():
        sheet_name = tname[:31]
        ws = wb.create_sheet(title=sheet_name)
        # 三张主表：在表格最前面重复列出主要合计/小计（摘要）
        disp_rows = rows
        top_names = SUMMARY_TOP.get(tname)
        if top_names:
            top = [r for n in top_names for r in rows if r['name'] == n]
            disp_rows = top + rows
        write_sheet(ws, tname, disp_rows, show_yoy=False)
        # 空一行后写同比
        ws.append([])
        write_sheet(ws, tname, disp_rows, show_yoy=True)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 26
        for col in range(2, 2 + len(years) * len(companies)):
            ws.column_dimensions[get_column_letter(col)].width = 14

    # 计算说明页：列出固定指标 / 杜邦分析每个指标的计算公式，便于 Excel 自解释
    note_rows = []
    for tname in ('固定指标', '杜邦分析'):
        for r in body['tables'].get(tname, []):
            if r.get('note'):
                g = r.get('group') or ('杜邦拆解' if tname == '杜邦分析' else '')
                note_rows.append((g, r['name'], r['note']))
    if note_rows:
        ws = wb.create_sheet(title='计算说明')
        ws.append(['指标分组', '指标名称', '计算公式 / 说明'])
        for g, n, note in note_rows:
            ws.append([g, n, note])
        ws.cell(1, 1).font = Font(bold=True, size=13, color='0C447C')
        for ci in (1, 2, 3):
            ws.cell(1, ci).fill = head_fill
            ws.cell(1, ci).font = bold_font
            ws.cell(1, ci).border = border
        for ri in range(2, 2 + len(note_rows)):
            for ci in (1, 2, 3):
                ws.cell(ri, ci).border = border
            ws.cell(ri, 3).alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 70

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 货币说明页：含港股时，明确标注各公司财报单位（港币 HKD / 人民币 CNY）
    has_hk = any(c.get('currency') == 'HKD' for c in companies)
    if has_hk:
        from openpyxl.utils import get_column_letter as _gcl
        ws = wb.create_sheet(title='说明')
        ws.append(['公司', '市场', '财报货币'])
        for c in companies:
            ws.append([c.get('name', ''), c.get('market', 'A股'),
                       '港币(HKD)' if c.get('currency') == 'HKD' else '人民币(CNY)'])
        ws.append([])
        ws.append(['注：港股财报数值单位为港币(HKD)，A股为人民币(CNY)；比率与同比(%)指标不受货币单位影响。'])
        ws.cell(1, 1).font = Font(bold=True, size=13, color='0C447C')
        for ci in (1, 2, 3):
            ws.cell(2, ci).fill = head_fill
            ws.cell(2, ci).font = bold_font
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 16
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
